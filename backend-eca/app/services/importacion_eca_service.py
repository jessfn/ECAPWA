"""Importación masiva de ECA — ECA-007.

Flujo en dos pasos, ambos síncronos:

1. `iniciar_importacion`: parsea el archivo (CSV o XLSX), detecta/valida la
   **columna identificador estable** (DP-2 — sin ella, se rechaza el
   archivo entero, sin insertar ni deduplicar por nombre/municipio), valida
   cada fila, y guarda un `LoteImportacion` en estado `VALIDADO` con las
   filas ya limpias y los errores en `resumen` (jsonb). Nada se escribe en
   `ecas` todavía.
2. `confirmar_importacion`: hace el **upsert transaccional por
   `clave_fuente`** de las filas válidas del lote. Idempotente: confirmar
   dos veces el mismo lote no duplica (la segunda vez es un no-op que
   devuelve el mismo resumen).

Columnas del CSV/XLSX esperadas (además de la columna identificador):
`nombre`, `estado_clave_inegi`, `municipio_clave_inegi`, y opcionalmente
`clave_institucional`, `localidad_nombre`, `latitud`, `longitud`.
"""
from __future__ import annotations

import csv
import io
import uuid as uuid_lib
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.audit import registrar_evento
from app.models.eca import Eca
from app.models.lote_importacion import LoteImportacion
from app.models.usuario import Usuario
from app.repositories import ecas as repo_ecas
from app.repositories import geo as repo_geo
from app.schemas.eca import ErrorFilaImportacion

COLUMNAS_IDENTIFICADOR_CANDIDATAS = ("clave_fuente", "id_eca", "folio", "clave", "clave_institucional")
COLUMNAS_REQUERIDAS = {"nombre", "estado_clave_inegi", "municipio_clave_inegi"}


class SinIdentificadorEstableError(ValueError):
    """DP-2: el archivo no trae (ni se indicó) una columna identificador estable."""


class LoteNoEncontradoError(Exception):
    pass


class LoteEnEstadoInvalidoError(Exception):
    pass


def _parsear_archivo(contenido: bytes, nombre_archivo: str) -> list[dict[str, str]]:
    nombre_lower = nombre_archivo.lower()
    if nombre_lower.endswith(".xlsx"):
        libro = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        hoja = libro.active
        filas_iter = hoja.iter_rows(values_only=True)
        encabezado = [str(c).strip() if c is not None else "" for c in next(filas_iter)]
        filas = []
        for fila in filas_iter:
            if all(v is None for v in fila):
                continue
            filas.append({encabezado[i]: ("" if v is None else str(v)) for i, v in enumerate(fila)})
        return filas

    texto = contenido.decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(texto)))


def _detectar_columna_identificador(columnas: set[str], sugerida: str | None) -> str | None:
    if sugerida:
        return sugerida if sugerida in columnas else None
    for candidata in COLUMNAS_IDENTIFICADOR_CANDIDATAS:
        if candidata in columnas:
            return candidata
    return None


def _parsear_decimal(valor: str | None) -> Decimal | None:
    if not valor or not valor.strip():
        return None
    try:
        return Decimal(valor.strip())
    except InvalidOperation:
        return None


def iniciar_importacion(
    db: Session,
    *,
    contenido: bytes,
    nombre_archivo: str,
    columna_identificador: str | None,
    actor: Usuario,
) -> LoteImportacion:
    filas_crudas = _parsear_archivo(contenido, nombre_archivo)
    if not filas_crudas:
        raise ValueError("El archivo no tiene filas.")

    columnas = set(filas_crudas[0].keys())
    col_id = _detectar_columna_identificador(columnas, columna_identificador)
    if col_id is None:
        # DP-2: sin identificador estable, se detiene por completo — nunca
        # se improvisa una deduplicación por nombre/municipio.
        raise SinIdentificadorEstableError(
            "El archivo no tiene una columna identificador estable (se buscó: "
            f"{', '.join(COLUMNAS_IDENTIFICADOR_CANDIDATAS)}) ni se indicó una explícitamente."
        )

    faltantes = COLUMNAS_REQUERIDAS - columnas
    if faltantes:
        raise ValueError(f"Faltan columnas requeridas en el archivo: {sorted(faltantes)}")

    estados_por_clave = {e.clave_inegi: e for e in repo_geo.listar_estados(db)}
    municipios_por_clave = {m.clave_inegi: m for m in repo_geo.listar_todos_municipios(db)}

    errores: list[ErrorFilaImportacion] = []
    filas_validas: list[dict] = []
    claves_fuente_en_archivo: set[str] = set()

    for indice, fila in enumerate(filas_crudas, start=2):  # fila 1 = encabezado
        clave_fuente = (fila.get(col_id) or "").strip()
        nombre = (fila.get("nombre") or "").strip()
        clave_estado = (fila.get("estado_clave_inegi") or "").strip()
        clave_municipio = (fila.get("municipio_clave_inegi") or "").strip()

        if not clave_fuente:
            errores.append(ErrorFilaImportacion(fila=indice, campo=col_id, mensaje="Identificador vacío."))
            continue
        if clave_fuente in claves_fuente_en_archivo:
            errores.append(
                ErrorFilaImportacion(
                    fila=indice, campo=col_id, mensaje=f"Identificador duplicado en el archivo: {clave_fuente}."
                )
            )
            continue
        if not nombre:
            errores.append(ErrorFilaImportacion(fila=indice, campo="nombre", mensaje="Nombre requerido."))
            continue
        estado = estados_por_clave.get(clave_estado)
        if estado is None:
            errores.append(
                ErrorFilaImportacion(
                    fila=indice, campo="estado_clave_inegi", mensaje=f"Estado desconocido: {clave_estado}."
                )
            )
            continue
        municipio = municipios_por_clave.get(clave_municipio)
        if municipio is None or municipio.estado_id != estado.id:
            errores.append(
                ErrorFilaImportacion(
                    fila=indice,
                    campo="municipio_clave_inegi",
                    mensaje=f"Municipio desconocido o no pertenece al estado: {clave_municipio}.",
                )
            )
            continue

        latitud = _parsear_decimal(fila.get("latitud"))
        longitud = _parsear_decimal(fila.get("longitud"))

        claves_fuente_en_archivo.add(clave_fuente)
        filas_validas.append(
            {
                "fila": indice,
                "clave_fuente": clave_fuente,
                "nombre": nombre,
                "estado_id": estado.id,
                "municipio_id": municipio.id,
                "clave_institucional": (fila.get("clave_institucional") or "").strip() or None,
                "localidad_nombre": (fila.get("localidad_nombre") or "").strip() or None,
                "latitud": str(latitud) if latitud is not None else None,
                "longitud": str(longitud) if longitud is not None else None,
            }
        )

    lote = LoteImportacion(
        tipo="ECA",
        archivo_nombre=nombre_archivo,
        total_filas=len(filas_crudas),
        filas_validas=len(filas_validas),
        filas_con_error=len(errores),
        estado="VALIDADO",
        resumen={
            "columna_identificador": col_id,
            "filas_validas_datos": filas_validas,
            "errores": [e.model_dump() for e in errores],
        },
    )
    repo_ecas.crear_lote(db, lote)
    db.commit()
    db.refresh(lote)
    return lote


def confirmar_importacion(db: Session, *, lote_uuid: uuid_lib.UUID, actor: Usuario) -> tuple[LoteImportacion, int, int]:
    lote = repo_ecas.obtener_lote_por_uuid(db, lote_uuid)
    if lote is None:
        raise LoteNoEncontradoError(f"Lote no encontrado: {lote_uuid}")

    if lote.estado == "CONFIRMADO":
        # Idempotente: reimportar/reconfirmar no debe duplicar ni fallar.
        resumen = lote.resumen or {}
        return lote, resumen.get("creadas", 0), resumen.get("actualizadas", 0)

    if lote.estado != "VALIDADO":
        raise LoteEnEstadoInvalidoError(f"El lote está en estado {lote.estado}, no se puede confirmar.")

    filas = lote.resumen.get("filas_validas_datos", [])
    creadas = actualizadas = 0

    for fila in filas:
        existente = repo_ecas.obtener_por_clave_fuente(db, fila["clave_fuente"])
        latitud = Decimal(fila["latitud"]) if fila.get("latitud") else None
        longitud = Decimal(fila["longitud"]) if fila.get("longitud") else None

        if existente is None:
            repo_ecas.crear_eca(
                db,
                Eca(
                    clave_fuente=fila["clave_fuente"],
                    clave_institucional=fila.get("clave_institucional"),
                    nombre=fila["nombre"],
                    estado_id=fila["estado_id"],
                    municipio_id=fila["municipio_id"],
                    localidad_nombre=fila.get("localidad_nombre"),
                    latitud=latitud,
                    longitud=longitud,
                    fuente_carga="IMPORTACION",
                    lote_importacion_id=lote.id,
                    creado_por=actor.id,
                ),
            )
            creadas += 1
        else:
            existente.nombre = fila["nombre"]
            existente.estado_id = fila["estado_id"]
            existente.municipio_id = fila["municipio_id"]
            existente.clave_institucional = fila.get("clave_institucional")
            existente.localidad_nombre = fila.get("localidad_nombre")
            existente.latitud = latitud
            existente.longitud = longitud
            existente.actualizado_por = actor.id
            db.add(existente)
            actualizadas += 1

    lote.estado = "CONFIRMADO"
    lote.confirmado_en = datetime.now(timezone.utc)
    lote.resumen = {**lote.resumen, "creadas": creadas, "actualizadas": actualizadas}
    db.add(lote)

    registrar_evento(
        db,
        accion="eca.importacion_confirmada",
        modulo="ecas",
        origen="IMPORTACION",
        actor_usuario_id=actor.id,
        entidad_tipo="lote_importacion",
        entidad_id=lote.id,
        entidad_uuid=lote.uuid,
        descripcion=f"Importación de ECA confirmada: {creadas} altas, {actualizadas} actualizaciones",
    )
    db.commit()
    db.refresh(lote)
    return lote, creadas, actualizadas
